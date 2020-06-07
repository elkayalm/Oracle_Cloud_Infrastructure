# Convert excel to csv
# Create new sheets
# Add sheets together and apply auto-filter and a column auto-fit
# 30-april-2020          Created         Mohamed Elkayal

import sys
import pandas as pd
from openpyxl import load_workbook

column_notation = ["A", "B", "C", "D", "E", "F", "G", "H", "I", "J", "K", "L", "M", "N", "O"]

def conversion(oci_file,psm_file):
    # Convert from excel to csv
    read_file2 = pd.read_csv (oci_file)
    read_file2.to_excel ("oci_resources.xlsx", index = None, header=True)

    read_file = pd.read_csv (psm_file)
    read_file.to_excel ("psm_resources.xlsx", index = None, header=True)


def create_sheets():

    workbook.create_sheet("psm_resources")
    workbook.create_sheet("oci_resources")

def copy_data(file_name, sheet_name):

    workbook1 = load_workbook(filename=file_name)
    worksheet = workbook[sheet_name]

    worksheet1 = workbook1.active

    # Get the maximum number of rows/columns
    max_row_count = worksheet1.max_row
    max_column_count = worksheet1.max_column

    column = 1

    while column != (max_row_count + 1):
        row = 1
        while row != (max_column_count + 1):
            pasted_cell = worksheet.cell(column,row)
            copied_cell =  worksheet1.cell(column,row).value

            # Assign the value from the first file to the original
            pasted_cell.value = copied_cell

            row += 1
        column += 1

    # Adding filters to all columns
    worksheet.auto_filter.ref = f"A1:{column_notation[(max_column_count - 1)]}1"

def column_autofit(sheetname):

    ws = workbook[sheetname]

    for i in column_notation:
        column_length = []

        # Check if the header exists or not
        if ws[f"{i}1"].value != None:
            for cell in ws[i]:
                try:
                    # Add the length of each column to the list
                    column_length.append(len(cell.value))
                except:
                    pass
            # Assign the max column length to be the column width
            ws.column_dimensions[i].width = max(column_length)


if __name__ == "__main__":
    # Get file names from command line
    if len(sys.argv) != 3:
        original_file = sys.argv[1]
        oci_file = sys.argv[2]
        psm_file = sys.argv[3]
    else:
        print("Please check the supplied parameters")

    conversion(oci_file, psm_file)

    workbook = load_workbook(filename=original_file)

    create_sheets()

    copy_data("psm_resources.xlsx", "psm_resources")
    copy_data("oci_resources.xlsx", "oci_resources")

    column_autofit("psm_resources")
    column_autofit("oci_resources")

    workbook.save(filename=original_file)
